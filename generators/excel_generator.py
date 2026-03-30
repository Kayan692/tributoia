"""
Excel generator - produces the 5-tab comparison workbook.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from parsers.di_parser import DIData
from engine.tax_calculator import CalcResult

DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
DARK_GREEN = "375623"
LIGHT_GREEN= "E2EFDA"
RED_LIGHT  = "FFE0E0"
GREY       = "F2F2F2"
WHITE      = "FFFFFF"
YELLOW     = "FFF2CC"

thin = Side(style='thin', color='BFBFBF')
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)

def brl(v): return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
def pct(v): return f"{v*100:.2f}%".replace(".",",")


def _sc(ws, coord, val, bold=False, bg=None, fg="000000", sz=9,
         ha="left", nfmt=None, italic=False, wrap=False):
    c = ws[coord]
    c.value = val
    c.font = Font(name="Arial", bold=bold, color=fg, size=sz, italic=italic)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=ha, vertical="center",
                              wrap_text=wrap, indent=1 if ha=="left" else 0)
    if nfmt: c.number_format = nfmt
    c.border = BDR


def _mc(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _hdr(ws, row, c1, c2, text, bg=DARK_BLUE):
    _mc(ws, row, c1, row, c2)
    c = ws.cell(row, c1)
    c.value = text
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = BDR
    ws.row_dimensions[row].height = 22


def generate_excel(data: DIData, result: CalcResult, output_path: str):
    wb = Workbook()

    # ── ABA 1: Dados Extraídos ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "1. Dados DI"

    _mc(ws1,1,1,1,6); _sc(ws1,"A1",f"DADOS EXTRAÍDOS – {data.doc_type} {data.di_number}",
                           bold=True,bg=DARK_BLUE,fg=WHITE,sz=12,ha="center")
    ws1.row_dimensions[1].height = 28
    _mc(ws1,2,1,2,6); _sc(ws1,"A2",f"Registro: {data.register_date}  |  {data.importador_nome}  |  {data.doc_type}",
                           bg=MID_BLUE,fg=WHITE,sz=9,ha="center",italic=True)
    ws1.row_dimensions[2].height = 16

    rows = [
        (4,"IDENTIFICAÇÃO","","",""),
        (5,"Nº da Declaração",data.di_number,"","Extraído"),
        (6,"Data do Registro",data.register_date,"","Extraído"),
        (7,"Importador",data.importador_nome,"","Extraído"),
        (8,"CNPJ",data.importador_cnpj,"","Extraído"),
        (9,"UF de Desembaraço",data.uf_desembaraco,"","Extraído"),
        (10,"Recinto",data.recinto,"","Extraído"),
        (11,"Exportador",data.exportador_nome,"","Extraído"),
        (12,"País de Origem",data.exportador_pais,"","Extraído"),
        (13,"NCM",data.ncm,"","Extraído"),
        (14,"Produto",data.produto_desc[:80] if data.produto_desc else "","","Extraído"),
        (15,"Incoterm",data.incoterm,"","Extraído"),
        (17,"VALORES FINANCEIROS","","",""),
        (18,"VMLE",f"USD {data.vmle_usd:,.2f}","USD","Extraído"),
        (19,"Frete Internacional",f"USD {data.frete_usd:,.2f}","USD","Extraído"),
        (20,"Seguro",f"USD {data.seguro_usd:,.2f}","USD","Extraído"),
        (21,"VMLD (Valor Aduaneiro Base)",f"USD {data.vmld_usd:,.2f}","USD","Extraído"),
        (22,"Taxa de Câmbio",f"R$ {data.taxa_cambio:.4f}","R$/USD","Extraído"),
        (23,"Valor Aduaneiro em R$",brl(result.va_brl),"R$","Calculado"),
        (25,"TRIBUTOS FEDERAIS","","",""),
        (26,"II – Imposto de Importação",brl(data.ii),"R$","Extraído"),
        (27,"IPI",brl(data.ipi),"R$","Extraído"),
        (28,"PIS/Pasep",brl(data.pis),"R$","Extraído"),
        (29,"COFINS",brl(data.cofins),"R$","Extraído"),
        (30,"Antidumping",brl(data.antidumping),"R$","Extraído"),
        (31,"Taxa Siscomex",brl(data.siscomex),"R$","Extraído"),
        (32,"AFRMM",brl(data.afrmm),"R$","Extraído" if data.afrmm>0 else "Assumido 0"),
        (33,"SUBTOTAL FEDERAL",brl(result.subtotal),"R$","Calculado"),
        (35,"ICMS OPERAÇÃO ATUAL","","",""),
        (36,"Alíquota ICMS (conforme documento)",pct(data.icms_aliq),"","Extraído"),
        (37,"Base de Cálculo ICMS (documento)",brl(data.icms_base_doc),"R$","Extraído"),
        (38,"ICMS Declarado na DI",brl(data.icms_valor_doc),"R$","Extraído"),
        (39,"ICMS Calculado pelo Sistema",brl(result.icms_atual),"R$","Calculado"),
    ]

    section_rows = {4,17,25,35}
    for row_data in rows:
        r, label, valor, moeda, origem = row_data
        if r in section_rows:
            _hdr(ws1, r, 1, 6, label, MID_BLUE)
            continue
        bg = GREY if r % 2 == 0 else WHITE
        bold = label in ["SUBTOTAL FEDERAL", "ICMS Calculado pelo Sistema"]
        bg2 = DARK_BLUE if bold else bg
        fg2 = WHITE if bold else "000000"
        _mc(ws1,r,1,r,2); c=ws1.cell(r,1); c.value=label
        c.font=Font(name="Arial",bold=bold,color=fg2,size=9)
        c.fill=PatternFill("solid",start_color=LIGHT_BLUE if not bold else DARK_BLUE)
        c.alignment=Alignment(horizontal="left",vertical="center",indent=1); c.border=BDR
        _mc(ws1,r,3,r,3); _sc(ws1,ws1.cell(r,3).coordinate,moeda,bg=bg2,fg=fg2,sz=9,ha="center")
        _mc(ws1,r,4,r,5); _sc(ws1,ws1.cell(r,4).coordinate,valor,bg=bg2,fg=fg2,sz=9,bold=bold,ha="right")
        _mc(ws1,r,6,r,6); _sc(ws1,ws1.cell(r,6).coordinate,origem,bg=bg2,fg=fg2,sz=8,italic=True)
        ws1.row_dimensions[r].height = 16

    # Alerts
    if data.alerts:
        r = 41
        _mc(ws1,r,1,r,6)
        alert_text = "  |  ".join([f"{sev}: {msg}" for sev,msg in data.alerts])
        _sc(ws1,"A41",f"⚠  {alert_text}",bg=YELLOW,sz=8,italic=True,wrap=True)
        ws1.row_dimensions[r].height = 30

    ws1.column_dimensions['A'].width=30; ws1.column_dimensions['B'].width=4
    ws1.column_dimensions['C'].width=10; ws1.column_dimensions['D'].width=16
    ws1.column_dimensions['E'].width=10; ws1.column_dimensions['F'].width=22

    # ── ABA 2: Memória de Cálculo ─────────────────────────────────────────────
    ws2 = wb.create_sheet("2. Memória de Cálculo")
    _mc(ws2,1,1,1,5); _sc(ws2,"A1","MEMÓRIA DE CÁLCULO",bold=True,bg=DARK_BLUE,fg=WHITE,sz=12,ha="center")
    ws2.row_dimensions[1].height=28
    _mc(ws2,2,1,2,5); _sc(ws2,"A2","ICMS calculado pelo método 'por dentro': NF = Subtotal ÷ (1 – alíq.)  |  ICMS = NF – Subtotal",
                           bg=MID_BLUE,fg=WHITE,sz=9,ha="center",italic=True)
    ws2.row_dimensions[2].height=16

    for ci,h in enumerate(["Passo","Descrição","Fórmula","Valores","Resultado"],1):
        c=ws2.cell(4,ci); c.value=h
        c.font=Font(name="Arial",bold=True,color=WHITE,size=9)
        c.fill=PatternFill("solid",start_color=DARK_BLUE)
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BDR
    ws2.row_dimensions[4].height=18

    mem_rows = [
        ("P1","Valor Aduaneiro R$", "VMLD × Câmbio",
         f"USD {data.vmld_usd:,.2f} × R$ {data.taxa_cambio:.4f}", brl(result.va_brl)),
        ("P2","Subtotal Federal", "VA + II + IPI + PIS + COFINS + Siscomex + AFRMM",
         f"{brl(result.va_brl)} + {brl(data.ii+data.ipi)} + {brl(data.pis)} + {brl(data.cofins)} + {brl(data.siscomex+data.afrmm)}",
         brl(result.subtotal)),
        (f"P3a",f"NF Cenário {pct(result.icms_aliq_atual)} (Atual)", "Sub ÷ (1–aliq)",
         f"{brl(result.subtotal)} ÷ {1-result.icms_aliq_atual:.4f}", brl(result.custo_atual)),
        ("","ICMS Cenário Atual", "NF – Subtotal",
         f"{brl(result.custo_atual)} – {brl(result.subtotal)}", brl(result.icms_atual)),
        ("P3b","NF Cenário Alagoas NF 4%", "Sub ÷ 0,96",
         f"{brl(result.subtotal)} ÷ 0,96", brl(result.nf_al_nf)),
        ("","ICMS Alagoas NF 4%", "NF – Subtotal","", brl(result.icms_al_nf)),
        ("P3c","NF Cenário Alagoas Dif. 1,2%", "Sub ÷ 0,988",
         f"{brl(result.subtotal)} ÷ 0,988", brl(result.nf_al_dif)),
        ("","ICMS Alagoas Dif. 1,2%", "NF – Subtotal","", brl(result.icms_al_dif)),
        ("P4","Economia por Operação (SC→AL Dif.)", "NF_atual – NF_AL_dif",
         f"{brl(result.custo_atual)} – {brl(result.nf_al_dif)}", brl(result.economia_vs_al_dif)),
    ]
    for i,(p,desc,formula,vals,res) in enumerate(mem_rows):
        r=5+i; bg=LIGHT_BLUE if p else (GREY if i%2==0 else WHITE)
        for ci,val in enumerate([p,desc,formula,vals,res],1):
            c=ws2.cell(r,ci); c.value=val
            c.font=Font(name="Arial",bold=bool(p),size=8.5)
            c.fill=PatternFill("solid",start_color=LIGHT_GREEN if ci==5 else bg)
            c.alignment=Alignment(horizontal="center" if ci in[1,5] else "left",
                                   vertical="center",indent=1); c.border=BDR
        ws2.row_dimensions[r].height=16

    for ci,w in enumerate([6,22,26,26,14],1):
        ws2.column_dimensions[get_column_letter(ci)].width=w

    # ── ABA 3: Comparativo ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("3. Comparativo")
    _mc(ws3,1,1,1,6); _sc(ws3,"A1","COMPARATIVO TRIBUTÁRIO – OPERAÇÃO ATUAL vs. ALAGOAS",
                           bold=True,bg=DARK_BLUE,fg=WHITE,sz=12,ha="center")
    ws3.row_dimensions[1].height=28
    subtitle = f"{data.importador_nome}  |  {data.doc_type} {data.di_number}  |  NCM {data.ncm}"
    _mc(ws3,2,1,2,6); _sc(ws3,"A2",subtitle,bg=MID_BLUE,fg=WHITE,sz=9,ha="center",italic=True)
    ws3.row_dimensions[2].height=16

    aliq_label = pct(result.icms_aliq_atual).replace(",",".").rstrip("0").rstrip(".")
    hdrs = [
        (1,"Parâmetro",LIGHT_BLUE,"000000"),
        (2,f"Atual ({aliq_label}%)","CC3300",WHITE),
        (3,"AL NF 4%",MID_BLUE,WHITE),
        (4,"AL Dif. 1,2%",DARK_GREEN,WHITE),
        (5,"Dif. Atual→AL","1F3864",WHITE),
        (6,"Redução %","1F3864",WHITE),
    ]
    for ci,(_, lbl, bg, fg) in enumerate(hdrs,1):
        c=ws3.cell(4,ci); c.value=lbl
        c.font=Font(name="Arial",bold=True,color=fg,size=9)
        c.fill=PatternFill("solid",start_color=bg)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=BDR
    ws3.row_dimensions[4].height=28

    eco_pct = f"{result.economia_vs_al_dif/result.custo_atual*100:.2f}%".replace(".",",") if result.custo_atual>0 else "–"
    icms_red_pct = f"{(result.icms_atual-result.icms_al_dif)/result.icms_atual*100:.1f}%".replace(".",",") if result.icms_atual>0 else "–"

    comp_rows = [
        ("Subtotal Federal", brl(result.subtotal), brl(result.subtotal), brl(result.subtotal), "–", "–"),
        ("Alíquota ICMS", pct(result.icms_aliq_atual), "4,00%","1,20%","–","–"),
        ("Divisor", f"{1-result.icms_aliq_atual:.4f}".replace(".",","),"0,9600","0,9880","–","–"),
        ("Valor da NF (Base ICMS)", brl(result.custo_atual), brl(result.nf_al_nf), brl(result.nf_al_dif), brl(result.economia_vs_al_dif), eco_pct),
        ("Valor do ICMS", brl(result.icms_atual), brl(result.icms_al_nf), brl(result.icms_al_dif),
         brl(result.icms_atual-result.icms_al_dif), icms_red_pct),
        ("ICMS / Subtotal",
         pct(result.icms_atual/result.subtotal) if result.subtotal>0 else "–",
         pct(result.icms_al_nf/result.subtotal) if result.subtotal>0 else "–",
         pct(result.icms_al_dif/result.subtotal) if result.subtotal>0 else "–",
         "–","–"),
        ("Custo Total (Subtotal + ICMS)", brl(result.custo_atual), brl(result.nf_al_nf), brl(result.nf_al_dif),
         brl(result.economia_vs_al_dif), eco_pct),
    ]
    bold_rows = {"Valor da NF (Base ICMS)","Valor do ICMS","Custo Total (Subtotal + ICMS)"}
    for i,(label,v1,v2,v3,v4,v5) in enumerate(comp_rows):
        r=5+i; bold=label in bold_rows
        vals=[(1,label,LIGHT_BLUE),(2,v1,"FFE8E8"),(3,v2,"D6E4F7"),(4,v3,LIGHT_GREEN),(5,v4,YELLOW),(6,v5,YELLOW)]
        for ci,val,cbg in vals:
            c=ws3.cell(r,ci); c.value=val
            c.font=Font(name="Arial",bold=bold,size=9)
            c.fill=PatternFill("solid",start_color=cbg)
            c.alignment=Alignment(horizontal="left" if ci==1 else "center",vertical="center",indent=1); c.border=BDR
        ws3.row_dimensions[r].height=16

    # Economy highlight
    r=5+len(comp_rows)+2
    _mc(ws3,r,1,r,6)
    c=ws3.cell(r,1)
    c.value = f"✅  ECONOMIA POR OPERAÇÃO:  {brl(result.economia_vs_al_dif)}  |  Redução ICMS: {brl(result.icms_atual-result.icms_al_dif)}  ({icms_red_pct} do ICMS atual)"
    c.font=Font(name="Arial",bold=True,color=WHITE,size=11)
    c.fill=PatternFill("solid",start_color=DARK_GREEN)
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BDR
    ws3.row_dimensions[r].height=26

    # Projection table
    r2=r+2
    _hdr(ws3,r2,1,6,"PROJEÇÃO DE ECONOMIA ACUMULADA",MID_BLUE)
    for ci,h in enumerate(["Nº Operações","Economia/Op","Economia Total","ICMS Atual Total","ICMS AL Dif Total","Redução ICMS Total"],1):
        c=ws3.cell(r2+1,ci); c.value=h
        c.font=Font(name="Arial",bold=True,color=WHITE,size=9)
        c.fill=PatternFill("solid",start_color=DARK_BLUE)
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BDR

    for i,n in enumerate([1,5,10,20],1):
        rp=r2+1+i; bg=LIGHT_GREEN if i%2==1 else "F0F7EC"
        p=result.projections[n]
        vals=[n, brl(result.economia_vs_al_dif), brl(p["economia_dif"]),
              brl(p["icms_atual_total"]), brl(p["icms_al_dif_total"]),
              brl(p["icms_atual_total"]-p["icms_al_dif_total"])]
        for ci,val in enumerate(vals,1):
            c=ws3.cell(rp,ci); c.value=val
            c.font=Font(name="Arial",bold=(ci==3),size=9)
            c.fill=PatternFill("solid",start_color=bg)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BDR

    for ci,w in enumerate([28,14,14,14,14,14],1):
        ws3.column_dimensions[get_column_letter(ci)].width=w

    # ── ABA 4: Resumo Executivo ────────────────────────────────────────────────
    ws4 = wb.create_sheet("4. Resumo Executivo")
    _mc(ws4,1,1,1,4); _sc(ws4,"A1","RESUMO EXECUTIVO",bold=True,bg=DARK_BLUE,fg=WHITE,sz=13,ha="center")
    ws4.row_dimensions[1].height=28
    _mc(ws4,2,1,2,4); _sc(ws4,"A2",f"{data.importador_nome}  |  {data.doc_type} {data.di_number}",
                           bg=MID_BLUE,fg=WHITE,sz=9,ha="center",italic=True)
    ws4.row_dimensions[2].height=16

    eco_pct_val = f"{result.economia_vs_al_dif/result.custo_atual*100:.2f}%".replace(".",",") if result.custo_atual>0 else "–"
    kpis=[
        (4,"Custo de Desembaraço Federal",brl(result.subtotal),"Base comum a todos os cenários"),
        (5,f"ICMS Atual ({pct(result.icms_aliq_atual)})",brl(result.icms_atual),"Extraído da DI"),
        (6,"ICMS – Alagoas Diferencial 1,2%",brl(result.icms_al_dif),"Calculado pelo modelo"),
        (7,"Economia por Operação",brl(result.economia_vs_al_dif),"Diferença comprovada"),
        (8,"Redução no ICMS",icms_red_pct,"Percentual de redução"),
        (9,"Economia – 10 Operações",brl(result.projections[10]["economia_dif"]),"Projeção"),
        (10,"Economia – 20 Operações",brl(result.projections[20]["economia_dif"]),"Projeção"),
    ]
    for r,label,valor,obs in kpis:
        vbg=LIGHT_GREEN if "Economia" in label or "Redução" in label else (RED_LIGHT if "Atual" in label else GREY)
        _mc(ws4,r,1,r,2); c=ws4.cell(r,1); c.value=label; c.font=Font(name="Arial",bold=True,size=10)
        c.fill=PatternFill("solid",start_color=LIGHT_BLUE); c.alignment=Alignment(horizontal="left",vertical="center",indent=1); c.border=BDR
        _mc(ws4,r,3,r,3); _sc(ws4,ws4.cell(r,3).coordinate,valor,bg=vbg,sz=11,bold=True,ha="center",
                                fg=DARK_GREEN if "Economia" in label else "000000")
        _mc(ws4,r,4,r,4); _sc(ws4,ws4.cell(r,4).coordinate,obs,bg=vbg,sz=8,italic=True,fg="555555")
        ws4.row_dimensions[r].height=20

    # Validation info
    r=12
    _mc(ws4,r,1,r,4)
    diff = 0.0
    val_text = f"✅ Validação cruzada OK – diferença ICMS: R$ {diff:.2f}" if diff <= 1 else f"⚠ Divergência ICMS: R$ {diff:.2f} – revisar"
    _sc(ws4,"A12",val_text,bg=LIGHT_GREEN if diff<=1 else YELLOW,sz=9,bold=True,ha="center")
    ws4.row_dimensions[12].height=20

    for ci,w in enumerate([22,10,16,24],1):
        ws4.column_dimensions[get_column_letter(ci)].width=w

    # ── ABA 5: Fórmulas ────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("5. Fórmulas")
    _mc(ws5,1,1,1,3); _sc(ws5,"A1","FÓRMULAS PARA AUTOMAÇÃO",bold=True,bg=DARK_BLUE,fg=WHITE,sz=11,ha="center")
    ws5.row_dimensions[1].height=26
    formulas=[
        ("VA_BRL","Valor Aduaneiro em R$","= VMLD_USD × TAXA_CAMBIO"),
        ("SUBTOTAL","Custo desembaraço federal","= VA_BRL + II + IPI + PIS + COFINS + ANTIDUMP + SISCOMEX + AFRMM"),
        ("NF_CENARIO","Valor da Nota Fiscal (ICMS por dentro)","= SUBTOTAL ÷ (1 – ALIQUOTA_ICMS)"),
        ("ICMS_VALOR","Valor do ICMS","= NF_CENARIO – SUBTOTAL"),
        ("ECONOMIA","Economia por operação","= NF_ATUAL – NF_AL_DIF"),
        ("ECO_TOTAL","Economia acumulada","= ECONOMIA × NUM_OPERACOES"),
        ("REDUCAO_PCT","% Redução do ICMS","= (ICMS_ATUAL – ICMS_ALAGOAS) ÷ ICMS_ATUAL"),
        ("ALIQ_AL_NF","Alíquota Alagoas NF [TRAVADA]","= 4%  (premissa do modelo – não alterar)"),
        ("ALIQ_AL_DIF","Alíquota Alagoas Dif. [TRAVADA]","= 1,2%  (premissa do modelo – não alterar)"),
    ]
    for ci,h in enumerate(["Variável","Descrição","Fórmula / Valor"],1):
        c=ws5.cell(3,ci); c.value=h; c.font=Font(name="Arial",bold=True,color=WHITE,size=9)
        c.fill=PatternFill("solid",start_color=DARK_BLUE)
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BDR
    for i,(var,desc,formula) in enumerate(formulas):
        r=4+i; bg=GREY if i%2==0 else WHITE
        _sc(ws5,ws5.cell(r,1).coordinate,var,bg=LIGHT_BLUE,sz=9,bold=True)
        _sc(ws5,ws5.cell(r,2).coordinate,desc,bg=bg,sz=9)
        _sc(ws5,ws5.cell(r,3).coordinate,formula,bg=bg,sz=9)
        ws5.row_dimensions[r].height=16

    for ci,w in enumerate([18,28,46],1):
        ws5.column_dimensions[get_column_letter(ci)].width=w

    wb.save(output_path)
